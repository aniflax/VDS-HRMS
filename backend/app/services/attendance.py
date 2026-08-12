from typing import Optional, List
from sqlalchemy.orm import Session
from sqlalchemy import and_, inspect
from sqlalchemy.exc import IntegrityError
from fastapi import HTTPException
from datetime import datetime, date, timedelta
from app.core.timezone import IST, get_local_now, get_local_today
from app.models.attendance import AttendanceLog, AttendanceStatus, AttendanceSource
from app.models.leave import LeaveRequest, LeaveRequestStatus, LeaveType
from app.models.sevak import Sevak, SevakStatusEnum, RoleEnum
from app.models.department import Department
from app.models.department_location import DepartmentLocation
from app.models.location import Location
from app.models.system_config import SystemConfig
from app.schemas.attendance import AttendanceMarkRequest, AttendanceManualUpdate
from app.utils.geo import calculate_distance
from app.services.attendance_realtime import broadcast_attendance_change
from app.services.week_off_history import get_effective_week_off_day

# Re-export so existing `from app.services.attendance import get_local_now`
# imports keep working.
__all__ = ["get_local_now", "get_local_today", "IST"]

TRACKED_ATTENDANCE_ROLES = [RoleEnum.SEVAK, RoleEnum.HOD]

def get_geo_threshold(db: Session) -> int:
    """Get geo threshold from system config, default to 500m."""
    config = db.query(SystemConfig).filter(SystemConfig.key == "GEO_THRESHOLD_METERS").first()
    if config:
        return int(config.value)
    return 500

def get_week_off_count(db: Session, sevak_id: str, week_start: date) -> int:
    """Get count of week-offs taken in the current week."""
    week_end = week_start + timedelta(days=6)
    count = db.query(AttendanceLog).filter(
        AttendanceLog.sevak_id == sevak_id,
        AttendanceLog.date >= week_start,
        AttendanceLog.date <= week_end,
        AttendanceLog.status == AttendanceStatus.WEEK_OFF
    ).count()
    return count

def get_week_start(d: date) -> date:
    """Get the start of the week (Monday) for a given date."""
    return d - timedelta(days=d.weekday())

def get_sevak_allocated_locations(db: Session, sevak_id: str) -> list:
    """Get all active locations allocated to a sevak."""
    from app.models.sevak_location import SevakLocation

    locations = db.query(SevakLocation).filter(
        SevakLocation.sevak_id == sevak_id,
        SevakLocation.is_active == True
    ).all()
    return locations

def find_closest_location(lat: float, lng: float, locations: list) -> tuple:
    """Find the closest location from a list of allocated locations. Returns (location, distance)."""
    if not locations:
        return None, None

    min_distance = float('inf')
    closest_loc = None

    for loc in locations:
        if loc.location_lat is not None and loc.location_lng is not None:
            dist = calculate_distance(lat, lng, loc.location_lat, loc.location_lng)
            if dist < min_distance:
                min_distance = dist
                closest_loc = loc

    return closest_loc, min_distance if closest_loc else None

def _table_exists(db: Session, table_name: str) -> bool:
    try:
        return inspect(db.get_bind()).has_table(table_name)
    except Exception:
        db.rollback()
        return False

def _candidate_location_name(location) -> str | None:
    return (
        getattr(location, "location_name", None)
        or getattr(location, "name", None)
    )

def _candidate_location_lat(location) -> float | None:
    return (
        getattr(location, "location_lat", None)
        if hasattr(location, "location_lat")
        else getattr(location, "latitude", None)
    )

def _candidate_location_lng(location) -> float | None:
    return (
        getattr(location, "location_lng", None)
        if hasattr(location, "location_lng")
        else getattr(location, "longitude", None)
    )

def _make_location_candidate(name: str | None, lat: float | None, lng: float | None, threshold_meters: int | None) -> dict | None:
    if lat is None or lng is None:
        return None
    return {
        "name": name,
        "lat": lat,
        "lng": lng,
        "threshold_meters": threshold_meters,
    }

def _normalize_threshold_meters(value, fallback: int = 500) -> int:
    try:
        if value is None or value == "":
            return fallback
        return int(float(value))
    except (TypeError, ValueError):
        return fallback

def get_attendance_location_candidates(db: Session, sevak: Sevak) -> list[dict]:
    """Return all office locations that can validate attendance for a sevak."""
    candidates = []
    default_threshold = get_geo_threshold(db)

    if _table_exists(db, "sevak_locations"):
        for location in get_sevak_allocated_locations(db, sevak.id):
            candidate = _make_location_candidate(
                _candidate_location_name(location),
                _candidate_location_lat(location),
                _candidate_location_lng(location),
                default_threshold,
            )
            if candidate:
                candidates.append(candidate)

    if sevak.department_id:
        if _table_exists(db, "locations") and _table_exists(db, "department_locations"):
            department_locations = db.query(Location).join(
                DepartmentLocation,
                DepartmentLocation.location_id == Location.id,
            ).filter(
                DepartmentLocation.department_id == sevak.department_id,
                Location.is_active == True,
            ).all()
            for location in department_locations:
                candidate = _make_location_candidate(
                    location.name,
                    location.latitude,
                    location.longitude,
                    location.geo_threshold_meters,
                )
                if candidate:
                    candidates.append(candidate)

        dept = db.query(Department).filter(Department.id == sevak.department_id).first()
        if dept and dept.location_lat is not None and dept.location_lng is not None:
            candidate = _make_location_candidate(
                dept.name,
                dept.location_lat,
                dept.location_lng,
                dept.geo_threshold_meters,
            )
            if candidate:
                candidates.append(candidate)

    deduped = []
    seen = set()
    for candidate in candidates:
        key = (
            candidate["name"],
            round(candidate["lat"], 7),
            round(candidate["lng"], 7),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(candidate)

    return deduped

def find_closest_attendance_location(db: Session, sevak: Sevak, lat: float, lng: float) -> tuple[dict | None, float | None]:
    closest_candidate = None
    closest_distance = None

    for candidate in get_attendance_location_candidates(db, sevak):
        distance = calculate_distance(lat, lng, candidate["lat"], candidate["lng"])
        if closest_distance is None or distance < closest_distance:
            closest_distance = distance
            closest_candidate = candidate

    return closest_candidate, closest_distance

def is_outside_attendance_location(db: Session, candidate: dict | None, distance: float | None) -> bool:
    if not candidate or distance is None:
        return False
    threshold = _normalize_threshold_meters(candidate.get("threshold_meters"), get_geo_threshold(db))
    return distance > threshold

def is_attendance_log_geo_mismatch(db: Session, sevak: Sevak, log: AttendanceLog) -> bool:
    """Return the effective geo mismatch state for a saved attendance log."""
    if log.geo_flagged:
        return True
    if log.location_lat is None or log.location_lng is None:
        return False

    try:
        closest_location, closest_distance = find_closest_attendance_location(
            db,
            sevak,
            log.location_lat,
            log.location_lng,
        )
    except Exception:
        db.rollback()
        return False

    return is_outside_attendance_location(db, closest_location, closest_distance)

def serialize_attendance_log(db: Session, log: AttendanceLog, sevak: Sevak) -> dict:
    data = {
        "id": log.id,
        "sevak_id": log.sevak_id,
        "date": log.date,
        "check_in_time": log.check_in_time,
        "check_out_time": log.check_out_time,
        "status": log.status,
        "source": log.source,
        "location_lat": log.location_lat,
        "location_lng": log.location_lng,
        "geo_flagged": log.geo_flagged,
        "is_manual": log.is_manual,
        "unlocked_by_id": log.unlocked_by_id,
        "location_name": None,
        "location_status": None,
        "location_map_url": None,
    }

    if log.location_lat is None or log.location_lng is None:
        data["location_status"] = "Not captured"
        return data

    data["location_map_url"] = f"https://www.google.com/maps?q={log.location_lat},{log.location_lng}"
    try:
        closest_location, closest_distance = find_closest_attendance_location(db, sevak, log.location_lat, log.location_lng)
    except Exception:
        db.rollback()
        closest_location, closest_distance = None, None
    location_mismatch = log.geo_flagged or is_outside_attendance_location(db, closest_location, closest_distance)
    data["geo_flagged"] = location_mismatch

    if closest_location:
        data["location_status"] = "Mismatch" if location_mismatch else "Verified"
    else:
        data["location_status"] = "Captured"

    return data

def get_day_leave_status(db: Session, sevak_id: str, target_date: date, default_week_off: str | None) -> tuple[bool, bool]:
    """Return whether target_date is an approved leave day or an effective week-off day."""
    week_start = get_week_start(target_date)
    week_end = week_start + timedelta(days=6)
    effective_week_off = get_effective_week_off_day(db, sevak_id, target_date, default_week_off)
    approved_week_off_requests = db.query(LeaveRequest).join(LeaveType, LeaveType.id == LeaveRequest.leave_type_id).filter(
        LeaveRequest.sevak_id == sevak_id,
        LeaveRequest.status == LeaveRequestStatus.APPROVED,
        LeaveType.name == "Week Off",
        LeaveRequest.start_date >= week_start,
        LeaveRequest.start_date <= week_end
    ).all()

    approved_requests = db.query(LeaveRequest).filter(
        LeaveRequest.sevak_id == sevak_id,
        LeaveRequest.status == LeaveRequestStatus.APPROVED,
        LeaveRequest.start_date <= target_date,
        LeaveRequest.end_date >= target_date
    ).all()

    if not approved_requests:
        is_default_week_off = (
            target_date.strftime("%A") == effective_week_off
            and not approved_week_off_requests
        )
        return False, is_default_week_off

    leave_type_map = {
        lt.id: lt.name
        for lt in db.query(LeaveType).filter(LeaveType.id.in_([req.leave_type_id for req in approved_requests])).all()
    }

    approved_week_off_this_week = bool(approved_week_off_requests)

    is_leave_day = any(leave_type_map.get(req.leave_type_id) != "Week Off" for req in approved_requests)
    is_default_week_off = (
        target_date.strftime("%A") == effective_week_off
        and not approved_week_off_this_week
    )
    is_approved_week_off_day = any(
        leave_type_map.get(req.leave_type_id) == "Week Off"
        for req in approved_requests
    )

    return is_leave_day, (is_default_week_off or is_approved_week_off_day)

def mark_attendance(db: Session, request: AttendanceMarkRequest, current_user: Sevak) -> AttendanceLog:
    """Mark attendance for the current day - single mark (no check-in/check-out)."""
    if current_user.role in [RoleEnum.SUPER_ADMIN, RoleEnum.ADMIN, RoleEnum.HR]:
        raise HTTPException(status_code=403, detail="Attendance is available only for Sevak and HOD accounts")

    now = get_local_now()
    today = now.date()

    if not current_user.activated_at or today < current_user.activated_at.date():
        raise HTTPException(status_code=400, detail="Attendance can be marked only from the account activation date")

    # Check if already marked today
    existing_log = db.query(AttendanceLog).filter(
        AttendanceLog.sevak_id == current_user.id,
        AttendanceLog.date == today
    ).first()

    if existing_log:
        raise HTTPException(status_code=400, detail="Attendance already marked for today")

    is_leave_day, is_week_off_day = get_day_leave_status(
        db=db,
        sevak_id=current_user.id,
        target_date=today,
        default_week_off=current_user.default_week_off
    )
    if is_leave_day:
        raise HTTPException(status_code=400, detail="Attendance cannot be marked on an approved leave day")
    if is_week_off_day:
        raise HTTPException(status_code=400, detail="Attendance cannot be marked on a week-off day")

    if request.lat is None or request.lng is None:
        raise HTTPException(status_code=400, detail="Location permission is required to mark attendance")

    closest_location, closest_distance = find_closest_attendance_location(db, current_user, request.lat, request.lng)
    geo_flagged = is_outside_attendance_location(db, closest_location, closest_distance)

    # Create new attendance record
    new_log = AttendanceLog(
        sevak_id=current_user.id,
        date=today,
        check_in_time=now,
        source=request.source,
        location_lat=request.lat,
        location_lng=request.lng,
        geo_flagged=geo_flagged,
        status=AttendanceStatus.PRESENT
    )
    db.add(new_log)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        existing_log = db.query(AttendanceLog).filter(
            AttendanceLog.sevak_id == current_user.id,
            AttendanceLog.date == today
        ).first()
        if existing_log:
            raise HTTPException(status_code=400, detail="Attendance already marked for today") from exc
        raise
    db.refresh(new_log)
    broadcast_attendance_change({
        "type": "attendance_marked",
        "sevak_id": current_user.id,
        "date": today.isoformat(),
        "source": request.source.value if hasattr(request.source, "value") else request.source,
    })
    return new_log

def mark_week_off(db: Session, target_date: date, current_user: Sevak) -> AttendanceLog:
    """Mark a day as week-off for the current user."""
    if current_user.role in [RoleEnum.SUPER_ADMIN, RoleEnum.ADMIN, RoleEnum.HR]:
        raise HTTPException(status_code=403, detail="Week off attendance is available only for Sevak and HOD accounts")

    if not current_user.activated_at or target_date < current_user.activated_at.date():
        raise HTTPException(status_code=400, detail="Week off can be marked only from the account activation date")

    # Check if already has a log for this date
    existing = db.query(AttendanceLog).filter(
        AttendanceLog.sevak_id == current_user.id,
        AttendanceLog.date == target_date
    ).first()

    if existing:
        raise HTTPException(status_code=400, detail="Attendance record already exists for this date")

    # Check week-off limit (1 per week)
    week_start = get_week_start(target_date)
    week_off_count = get_week_off_count(db, current_user.id, week_start)

    if week_off_count >= 1:
        raise HTTPException(status_code=400, detail="You have already taken a week-off this week")

    # Check if date is in the past or today
    today = get_local_today()
    if target_date > today:
        raise HTTPException(status_code=400, detail="Cannot mark week-off for future dates")

    # Check if date is in the current week
    if target_date < week_start:
        raise HTTPException(status_code=400, detail="Cannot mark week-off for previous weeks")

    new_log = AttendanceLog(
        sevak_id=current_user.id,
        date=target_date,
        status=AttendanceStatus.WEEK_OFF,
        source=AttendanceSource.WEB
    )
    db.add(new_log)
    db.commit()
    db.refresh(new_log)
    return new_log

def get_weekly_flagged_attendance(db: Session) -> list[AttendanceLog]:
    """HR Report fetcher for GEO_MISMATCH."""
    from datetime import timedelta
    last_week = get_local_today() - timedelta(days=7)

    logs = db.query(AttendanceLog).filter(
        AttendanceLog.geo_flagged == True,
        AttendanceLog.date >= last_week
    ).all()
    sevaks = {
        s.id: s
        for s in db.query(Sevak).filter(
            Sevak.id.in_([log.sevak_id for log in logs]),
            Sevak.role.in_(TRACKED_ATTENDANCE_ROLES),
        ).all()
    }
    return [log for log in logs if sevaks.get(log.sevak_id) and sevaks[log.sevak_id].activated_at and log.date >= sevaks[log.sevak_id].activated_at.date()]

def get_weekly_attendance(db: Session) -> list[AttendanceLog]:
    """HR Report fetcher for ALL weekly records."""
    from datetime import timedelta
    last_week = get_local_today() - timedelta(days=7)

    logs = db.query(AttendanceLog).filter(
        AttendanceLog.date >= last_week
    ).all()
    sevaks = {
        s.id: s
        for s in db.query(Sevak).filter(
            Sevak.id.in_([log.sevak_id for log in logs]),
            Sevak.role.in_(TRACKED_ATTENDANCE_ROLES),
        ).all()
    }
    return [log for log in logs if sevaks.get(log.sevak_id) and sevaks[log.sevak_id].activated_at and log.date >= sevaks[log.sevak_id].activated_at.date()]

def manual_update_attendance(db: Session, request: AttendanceManualUpdate, current_user: Sevak) -> AttendanceLog:
    """Manually update or create attendance record for a Sevak."""
    target_sevak = db.query(Sevak).filter(Sevak.id == request.sevak_id).first()
    if not target_sevak:
        raise HTTPException(status_code=404, detail="Sevak not found")
    if target_sevak.role not in TRACKED_ATTENDANCE_ROLES:
        raise HTTPException(status_code=400, detail="Attendance can be updated only for Sevak and HOD accounts")
    if not target_sevak.activated_at or request.date < target_sevak.activated_at.date():
        raise HTTPException(status_code=400, detail="Attendance can be updated only from the account activation date")

    # Find existing log
    log = db.query(AttendanceLog).filter(
        AttendanceLog.sevak_id == request.sevak_id,
        AttendanceLog.date == request.date
    ).first()

    if log:
        # Update existing
        log.status = request.status
        log.source = request.source
        log.is_manual = True
        log.unlocked_by_id = current_user.id
        log.geo_flagged = False  # Reset geo flag if manually updated
        if request.check_in_time:
            log.check_in_time = request.check_in_time
    else:
        # Create new
        log = AttendanceLog(
            sevak_id=request.sevak_id,
            date=request.date,
            check_in_time=request.check_in_time or get_local_now(),
            status=request.status,
            source=request.source,
            is_manual=True,
            unlocked_by_id=current_user.id,
            geo_flagged=False
        )
        db.add(log)

    db.commit()
    db.refresh(log)
    broadcast_attendance_change({
        "type": "attendance_updated",
        "sevak_id": request.sevak_id,
        "date": request.date.isoformat(),
        "source": request.source.value if hasattr(request.source, "value") else request.source,
    })
    return log

def get_sevaks_without_attendance_today(db: Session) -> list[Sevak]:
    """Get all active sevaks who haven't marked attendance today."""
    today = get_local_today()

    # Get all active sevaks
    all_sevaks = db.query(Sevak).filter(
        Sevak.is_active == True,
        Sevak.role.in_(TRACKED_ATTENDANCE_ROLES),
    ).all()

    # Get sevaks who have marked attendance today
    marked_today = db.query(AttendanceLog.sevak_id).filter(
        AttendanceLog.date == today
    ).all()
    marked_ids = {r[0] for r in marked_today}

    # Return sevaks who haven't marked and whose attendance period has started.
    return [
        s for s in all_sevaks
        if s.id not in marked_ids and s.activated_at and today >= s.activated_at.date()
    ]

def is_attendance_reminder_enabled(db: Session) -> bool:
    """Check if attendance reminder is enabled."""
    config = db.query(SystemConfig).filter(SystemConfig.key == "ATTENDANCE_REMINDER_ENABLED").first()
    if config:
        return config.value.lower() == "true"
    return False

def set_attendance_reminder(db: Session, enabled: bool, user_id: str):
    """Enable or disable attendance reminder."""
    config = db.query(SystemConfig).filter(SystemConfig.key == "ATTENDANCE_REMINDER_ENABLED").first()
    if config:
        config.value = str(enabled).lower()
        config.modified_by = user_id
    else:
        config = SystemConfig(
            key="ATTENDANCE_REMINDER_ENABLED",
            value=str(enabled).lower(),
            description="Enable/disable attendance reminder emails at 10:30 AM",
            access_level="SUPER_ADMIN",
            modified_by=user_id
        )
        db.add(config)
    db.commit()

def get_attendance_deadline(db: Session) -> str:
    """Get attendance deadline time (default 10:30 AM)."""
    config = db.query(SystemConfig).filter(SystemConfig.key == "ATTENDANCE_DEADLINE_TIME").first()
    if config:
        return config.value
    return "10:30"

def get_attendance_reminder_last_sent(db: Session) -> Optional[str]:
    """Get the date the last attendance reminder was sent."""
    config = db.query(SystemConfig).filter(SystemConfig.key == "ATTENDANCE_REMINDER_LAST_SENT_DATE").first()
    return config.value if config else None

def get_non_compliant_attendance(db: Session, start_date: date, end_date: date) -> list:
    """
    Get detailed attendance exception records for a date range.
    """
    results = []
    active_sevaks = db.query(Sevak).filter(
        Sevak.status == SevakStatusEnum.ACTIVE,
        Sevak.role.in_(TRACKED_ATTENDANCE_ROLES),
    ).all()

    # Get all logs and leaves in range for efficiency
    attendance_logs = db.query(AttendanceLog).filter(
        AttendanceLog.date >= start_date,
        AttendanceLog.date <= end_date
    ).all()
    attendance_map = {(log.sevak_id, log.date): log for log in attendance_logs}

    current_date = start_date
    while current_date <= end_date:
        # Weekend logic handled by Sevak default week off mainly, but here we assume Sat/Sun
        # are non-working for "Non-compliance" purposes unless marked.
        if current_date.weekday() >= 5:
            current_date += timedelta(days=1)
            continue

        for sevak in active_sevaks:
            if not sevak.activated_at or current_date < sevak.activated_at.date():
                continue
            log = attendance_map.get((sevak.id, current_date))

            # Check leave/weekoff status for this day
            is_leave, is_week_off = get_day_leave_status(db, sevak.id, current_date, sevak.default_week_off)

            if is_leave or is_week_off:
                continue

            if not log:
                # No attendance marked - ABSENT
                results.append({
                    'sevak_db_id': sevak.id,
                    'sevak_id': sevak.sevak_id,
                    'sevak_name': f"{sevak.first_name} {sevak.last_name}",
                    'date': current_date.strftime('%Y-%m-%d'),
                    'day': current_date.strftime('%A'),
                    'status': 'ABSENT',
                    'type': 'Missed Attendance',
                    'check_in_time': None,
                    'distance': None,
                    'remarks': 'No attendance marked'
                })
            elif log.status == AttendanceStatus.ABSENT:
                results.append({
                    'sevak_db_id': sevak.id,
                    'sevak_id': sevak.sevak_id,
                    'sevak_name': f"{sevak.first_name} {sevak.last_name}",
                    'date': current_date.strftime('%Y-%m-%d'),
                    'day': current_date.strftime('%A'),
                    'status': 'ABSENT',
                    'type': 'Marked Absent',
                    'check_in_time': None,
                    'distance': None,
                    'remarks': 'Marked as absent'
                })
            elif is_attendance_log_geo_mismatch(db, sevak, log):
                distance_str = f"{log.location_lat:.6f}, {log.location_lng:.6f}" if log.location_lat else 'N/A'
                results.append({
                    'sevak_db_id': sevak.id,
                    'sevak_id': sevak.sevak_id,
                    'sevak_name': f"{sevak.first_name} {sevak.last_name}",
                    'date': current_date.strftime('%Y-%m-%d'),
                    'day': current_date.strftime('%A'),
                    'status': log.status.value if hasattr(log.status, 'value') else log.status,
                    'type': 'Geo Mismatch',
                    'check_in_time': log.check_in_time.strftime('%H:%M') if log.check_in_time else None,
                    'distance': distance_str,
                    'remarks': 'Attendance marked outside geo threshold'
                })

        current_date += timedelta(days=1)
    return results

def get_monthly_aggregated_report(db: Session, year: int, month: int) -> list:
    """Get aggregated monthly stats (Working Days, Week Offs, Leaves, Attendance) per Sevak."""
    import calendar
    from app.models.department import Department

    start_date = date(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    end_date = date(year, month, last_day)
    total_days_in_month = last_day

    active_sevaks = db.query(Sevak).filter(
        Sevak.status == SevakStatusEnum.ACTIVE,
        Sevak.role.in_(TRACKED_ATTENDANCE_ROLES),
    ).all()
    departments = {d.id: d.name for d in db.query(Department).all()}

    # Get all active leave types (Dynamic categories)
    leave_types = db.query(LeaveType).filter(LeaveType.is_active == True).all()
    leave_type_names = [lt.name for lt in leave_types if lt.name != "Week Off"]

    # Pre-fetch all logs and leaves for the month
    attendance_logs = db.query(AttendanceLog).filter(
        AttendanceLog.date >= start_date,
        AttendanceLog.date <= end_date
    ).all()

    leave_requests = db.query(LeaveRequest).join(LeaveType).filter(
        LeaveRequest.status == LeaveRequestStatus.APPROVED,
        LeaveRequest.start_date <= end_date,
        LeaveRequest.end_date >= start_date
    ).add_columns(LeaveType.name.label("type_name")).all()

    results = []

    for sevak in active_sevaks:
        s_logs = [log for log in attendance_logs if log.sevak_id == sevak.id]
        s_leaves = [lr for lr in leave_requests if lr[0].sevak_id == sevak.id]

        # Attendance logs are the source of truth; geo status is recalculated from stored coordinates
        # so reports match the attendance calendar when location rules change after a log was created.
        geo_mismatch_by_log_id = {
            log.id: is_attendance_log_geo_mismatch(db, sevak, log)
            for log in s_logs
        }
        present_count = sum(
            1 for log in s_logs
            if log.status == AttendanceStatus.PRESENT and not geo_mismatch_by_log_id.get(log.id, False)
        )
        geo_mismatch_count = sum(1 for is_mismatch in geo_mismatch_by_log_id.values() if is_mismatch)

        leave_count = 0
        absent_count = 0
        week_off_count = 0

        # Initialize dynamic breakdown
        leave_breakdown = {name: 0 for name in leave_type_names}

        active_days_in_period = 0
        current_date = start_date
        while current_date <= end_date:
            if not sevak.activated_at or current_date < sevak.activated_at.date():
                current_date += timedelta(days=1)
                continue

            active_days_in_period += 1
            is_leave, is_week_off = get_day_leave_status(db, sevak.id, current_date, sevak.default_week_off)

            if is_week_off:
                week_off_count += 1
            elif is_leave:
                leave_count += 1
                # Identify which leave type
                for lr, tname in s_leaves:
                    if lr.start_date <= current_date <= lr.end_date:
                        if tname in leave_breakdown:
                            leave_breakdown[tname] += 1
                        break
            else:
                has_mark = any(log.date == current_date and log.status == AttendanceStatus.PRESENT for log in s_logs)
                if not has_mark:
                    absent_count += 1

            current_date += timedelta(days=1)

        results.append({
            'sevak_db_id': sevak.id,
            'sevak_id': sevak.sevak_id,
            'name': f"{sevak.first_name} {sevak.last_name}",
            'department_name': departments.get(sevak.department_id, 'Unassigned'),
            'total_days': active_days_in_period,
            'working_days': active_days_in_period - week_off_count,
            'week_off': week_off_count,
            'present': present_count,
            'leave': leave_count,
            'absent_days': absent_count,
            'geo_mismatch': geo_mismatch_count,
            'leave_breakdown': leave_breakdown,
            'leave_type_list': leave_type_names
        })

    return results

def generate_non_compliant_excel(db: Session, start_date: date, end_date: date) -> bytes:
    """Generate a two-sheet professional Excel report with dynamic leave columns and customized headers."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    month_name = start_date.strftime('%B %Y')
    data = get_monthly_aggregated_report(db, start_date.year, start_date.month)
    if not data:
        return b""

    leave_columns = data[0]['leave_type_list']
    num_leave_cols = len(leave_columns)

    wb = Workbook()

    # Styles
    dark_blue_fill = PatternFill(start_color="1B3B6F", end_color="1B3B6F", fill_type="solid")
    light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    # SHEET 1: Aggregated Summary
    ws1 = wb.active
    ws1.title = "Monthly Summary"

    ws1.merge_cells('A1:E2')
    title_cell = ws1['A1']
    title_cell.value = month_name
    title_cell.fill = dark_blue_fill
    title_cell.font = white_font
    title_cell.alignment = center_align

    ws1.append([])
    headers1 = ['Name', 'Present Days', 'Week Off Count', 'Leave Days', 'Absent Days']
    ws1.append(headers1)
    for cell in ws1[4]:
        cell.font = header_font
        cell.fill = light_blue_fill
        cell.border = border
        cell.alignment = center_align

    for row in data:
        ws1.append([row['name'], row['present'], row['week_off'], row['leave'], row['absent_days']])
        for cell in ws1[ws1.max_row]:
            cell.border = border

    for col in range(1, 6):
        ws1.column_dimensions[get_column_letter(col)].width = 22

    # SHEET 2: Detailed Summary
    ws2 = wb.create_sheet(title="Detailed Summary")

    # Calculation of last column
    # Dept(1) + Name(2) + Working(3) + Weekly(4) + Attendance(5,6) + Dynamic(7+) + Total Count + Absence
    last_col_idx = 6 + num_leave_cols + 2
    last_col_letter = get_column_letter(last_col_idx)

    ws2.merge_cells(f'A1:{last_col_letter}2')
    title_cell2 = ws2['A1']
    title_cell2.value = month_name
    title_cell2.fill = dark_blue_fill
    title_cell2.font = white_font
    title_cell2.alignment = center_align

    # Multi-level header merges
    ws2.merge_cells('A3:A4')
    ws2.cell(row=3, column=1, value='Department')
    ws2.merge_cells('B3:B4')
    ws2.cell(row=3, column=2, value='Name')
    ws2.merge_cells('C3:C4')
    ws2.cell(row=3, column=3, value='Working days')
    ws2.merge_cells('D3:D4')
    ws2.cell(row=3, column=4, value='Weekly offs')

    ws2.merge_cells('E3:F3')
    ws2.cell(row=3, column=5, value='Attendance')

    # Dynamic Merges for Leaves
    leaves_group_start = 7
    leaves_group_end = 7 + num_leave_cols # This col will be "Total Count"
    ws2.merge_cells(start_row=3, start_column=leaves_group_start, end_row=3, end_column=leaves_group_end)
    ws2.cell(row=3, column=leaves_group_start, value='Leaves')

    # Merge vertical for Absence
    absent_col_idx = leaves_group_end + 1
    ws2.merge_cells(start_row=3, start_column=absent_col_idx, end_row=4, end_column=absent_col_idx)
    ws2.cell(row=3, column=absent_col_idx, value='Days of Absence')

    # Sub-headers (Row 4)
    ws2.cell(row=4, column=5, value='Marked')
    ws2.cell(row=4, column=6, value='Not Marked')
    for i, name in enumerate(leave_columns):
        ws2.cell(row=4, column=leaves_group_start + i, value=name)
    ws2.cell(row=4, column=leaves_group_end, value='Total Count')

    # Styles for headers
    for r in range(3, 5):
        for c in range(1, last_col_idx + 1):
            cell = ws2.cell(row=r, column=c)
            cell.fill = dark_blue_fill
            cell.font = Font(color="FFFFFF", bold=True, size=10)
            cell.alignment = center_align
            cell.border = border

    # Data Rows
    for row in data:
        data_row = [
            row['department_name'],
            row['name'],
            row['working_days'],
            row['week_off'],
            row['present'],
            row['absent_days']
        ]
        # Dynamic leave breakdown
        for name in leave_columns:
            data_row.append(row['leave_breakdown'].get(name, 0))

        data_row.append(row['leave']) # Total Count
        data_row.append(row['absent_days']) # Days of Absence

        ws2.append(data_row)
        for cell in ws2[ws2.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ws2.cell(row=ws2.max_row, column=2).alignment = Alignment(horizontal='left')

    for col in range(1, last_col_idx + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 15
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['A'].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def get_non_compliant_summary(db: Session, start_date: date, end_date: date) -> dict:
    """Get summary statistics for attendance exceptions."""
    monthly_rows = get_monthly_aggregated_report(db, start_date.year, start_date.month)

    missed_attendance = sum(row['absent_days'] for row in monthly_rows)
    geo_mismatch = sum(row['geo_mismatch'] for row in monthly_rows)
    total_records = missed_attendance + geo_mismatch
    unique_sevaks = sum(
        1 for row in monthly_rows
        if row['absent_days'] > 0 or row['geo_mismatch'] > 0
    )

    return {
        'total_records': total_records,
        'missed_attendance': missed_attendance,
        'geo_mismatch': geo_mismatch,
        'unique_sevaks': unique_sevaks,
        'period': f"{start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}"
    }
